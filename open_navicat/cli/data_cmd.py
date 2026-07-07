"""Data operations CLI commands — browse, export, import, sync, generate."""

from __future__ import annotations

import typer
from rich.console import Console

from open_navicat.services.connection_manager import connection_manager
from open_navicat.services.query_engine import query_engine
from open_navicat.utils.output_formatter import format_output

data_app = typer.Typer(name="data", help="Browse, export, import & sync data", no_args_is_help=True)
console = Console()


def _get_active_conn() -> str:
    active = connection_manager.active_ids
    if not active:
        console.print("[red]No active connection. Use 'opennavicat conn open <name>' first.[/red]")
        raise typer.Exit(1)
    return active[0]


def _resolve_conn(conn_name: str) -> str:
    if conn_name:
        connections = connection_manager.list_saved()
        target = next((c for c in connections if c.name == conn_name), None)
        if not target:
            console.print(f"[red]Connection '{conn_name}' not found.[/red]")
            raise typer.Exit(1)
        connection_manager.connect(target)
        return target.id
    return _get_active_conn()


@data_app.command("browse")
def browse_table(
    table: str = typer.Argument(..., help="Table name (format: db.table)"),
    conn: str = typer.Option("", "--conn", "-c", help="Connection name"),
    limit: int = typer.Option(100, "--limit", "-l", help="Maximum rows"),
    offset: int = typer.Option(0, "--offset", "-o", help="Row offset"),
    where: str = typer.Option("", "--where", "-w", help="WHERE clause (SQL)"),
    order: str = typer.Option("", "--order", help="ORDER BY clause"),
    format: str = typer.Option("table", "--format", "-f", help="Output format: table|json|csv"),
) -> None:
    """Browse table data with pagination and filtering."""
    cid = _resolve_conn(conn)

    if "." not in table:
        console.print("[red]Use format: db.table (e.g. mydb.users)[/red]")
        raise typer.Exit(1)
    database, table_name = table.split(".", 1)

    # Detect engine for quoting
    from open_navicat.dal.connection_pool import connection_pool as _pool
    _conn = _pool.get(cid)
    _info = getattr(_conn, "_info", None) if _conn else None
    _engine = getattr(_info, "engine", "mysql") if _info else "mysql"
    q = '"' if _engine == "postgresql" else "`"

    # Build query
    parts = [f"SELECT * FROM {q}{database}{q}.{q}{table_name}{q}"]
    if where:
        parts.append(f"WHERE {where}")
    if order:
        parts.append(f"ORDER BY {order}")
    if _engine == "postgresql":
        parts.append(f"LIMIT {limit} OFFSET {offset}")
    else:
        parts.append(f"LIMIT {limit} OFFSET {offset}")

    sql = " ".join(parts)
    result = query_engine.execute(cid, sql)

    if not result.success:
        console.print(f"[red]Error: {result.error_message}[/red]")
        raise typer.Exit(1)

    if result.is_select:
        rows = [{c.name: str(v) if v is not None else None for c, v in zip(result.columns, row)} for row in result.rows]
        format_output(rows, format, title=f"{table} ({result.row_count} rows)")
    else:
        console.print(f"[green]{result.affected_rows} rows affected[/green]")


@data_app.command("export")
def export_data(
    table: str = typer.Argument(..., help="Table name (format: db.table)"),
    output: str = typer.Option("", "--output", "-o", help="Output file path"),
    format: str = typer.Option("csv", "--format", "-f", help="Export format: csv|json|excel"),
    conn: str = typer.Option("", "--conn", "-c", help="Connection name"),
    where: str = typer.Option("", "--where", "-w", help="WHERE clause"),
    limit: int = typer.Option(0, "--limit", "-l", help="Maximum rows (0 = all)"),
) -> None:
    """Export table data to file."""
    cid = _resolve_conn(conn)
    if "." not in table:
        console.print("[red]Use format: db.table (e.g. mydb.users)[/red]")
        raise typer.Exit(1)

    database, table_name = table.split(".", 1)
    parts = [f"SELECT * FROM `{database}`.`{table_name}`"]
    if where:
        parts.append(f"WHERE {where}")
    if limit:
        parts.append(f"LIMIT {limit}")

    result = query_engine.execute(cid, " ".join(parts))
    if not result.success:
        console.print(f"[red]Error: {result.error_message}[/red]")
        raise typer.Exit(1)

    if not result.is_select or not result.rows:
        console.print("[yellow]No data to export.[/yellow]")
        raise typer.Exit()

    # Determine output path
    if not output:
        ext = {"csv": ".csv", "json": ".json", "excel": ".xlsx"}[format]
        output = f"{table_name}{ext}"

    # Convert rows
    data = [{c.name: v for c, v in zip(result.columns, row)} for row in result.rows]

    if format == "json":
        import json
        with open(output, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    elif format == "csv":
        import csv
        with open(output, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=[c.name for c in result.columns])
            writer.writeheader()
            writer.writerows(data)
    elif format == "excel":
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = table_name
            ws.append([c.name for c in result.columns])
            for row in data:
                ws.append([row[c.name] for c in result.columns])
            wb.save(output)
        except ImportError:
            console.print("[red]openpyxl not installed. Install: pip install openpyxl[/red]")
            raise typer.Exit(1)

    console.print(f"[green]✓ Exported {len(data)} rows to {output}[/green]")


@data_app.command("import")
def import_data(
    table: str = typer.Argument(..., help="Target table (format: db.table)"),
    file: str = typer.Argument(..., help="Input file path"),
    format: str = typer.Option("auto", "--format", "-f", help="File format: csv|json|excel|auto"),
    conn: str = typer.Option("", "--conn", "-c", help="Connection name"),
    batch_size: int = typer.Option(500, "--batch-size", "-b", help="Rows per batch insert"),
) -> None:
    """Import data from file into a table."""
    cid = _resolve_conn(conn)
    if "." not in table:
        console.print("[red]Use format: db.table (e.g. mydb.users)[/red]")
        raise typer.Exit(1)
    database, table_name = table.split(".", 1)

    # Auto-detect format
    if format == "auto":
        ext = file.rsplit(".", 1)[-1].lower()
        format = {"json": "json", "csv": "csv", "xlsx": "excel"}.get(ext, "csv")

    # Load data
    data: list[dict] = []
    if format == "json":
        import json
        with open(file, "r", encoding="utf-8") as f:
            data = json.load(f)
    elif format == "csv":
        import csv
        with open(file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            data = list(reader)
    elif format == "excel":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(headers, row)))
        except ImportError:
            console.print("[red]openpyxl not installed.[/red]")
            raise typer.Exit(1)

    if not data:
        console.print("[yellow]No data found in file.[/yellow]")
        raise typer.Exit()

    # Batch insert
    total = 0
    from open_navicat.dal.connection_pool import connection_pool
    connector = connection_pool.get(cid)
    if not connector:
        console.print("[red]Connection lost.[/red]")
        raise typer.Exit(1)

    from open_navicat.dal.connection_pool import _loop as pool_loop
    from open_navicat.dal.connection_pool import connection_pool

    for i in range(0, len(data), batch_size):
        batch = data[i:i + batch_size]
        count = pool_loop.run_until_complete(connector.batch_insert(database, table_name, batch))
        total += count
        console.print(f"  [dim]Imported {total}/{len(data)} rows...[/dim]")

    console.print(f"[green]✓ Imported {total} rows into {table}[/green]")


@data_app.command("generate")
def generate_data(
    table: str = typer.Argument(..., help="Target table (format: db.table)"),
    count: int = typer.Option(100, "--count", "-n", help="Number of rows to generate"),
    conn: str = typer.Option("", "--conn", "-c", help="Connection name"),
    prompt: str = typer.Option("", "--prompt", "-p", help="AI prompt: describe realistic data rules"),
    preview: bool = typer.Option(True, "--preview/--yes", help="Preview before inserting"),
) -> None:
    """AI-powered test data generation — describe business rules, get realistic data."""
    cid = _resolve_conn(conn)
    if "." not in table:
        console.print("[red]Use format: db.table (e.g. mydb.users)[/red]")
        raise typer.Exit(1)
    database, table_name = table.split(".", 1)

    from open_navicat.services.ai_service import ai_service
    from open_navicat.services.metadata_service import metadata_service

    info = metadata_service.get_table_info(cid, database, table_name)
    if not info:
        console.print(f"[red]Table '{table}' not found.[/red]")
        raise typer.Exit(1)

    console.print(f"[yellow]🤖 AI generating {count} realistic rows for {table}...[/yellow]")
    rows = ai_service.generate_data(info, count, prompt)

    if not rows:
        console.print("[red]Failed to generate test data.[/red]")
        raise typer.Exit(1)

    if preview:
        format_output(rows[:5], "table", title=f"Preview: {len(rows)} rows to insert")
        confirm = typer.confirm(f"Insert {len(rows)} rows?")
        if not confirm:
            console.print("[yellow]Cancelled.[/yellow]")
            raise typer.Exit()

    from open_navicat.dal.connection_pool import _loop as pool_loop
    from open_navicat.dal.connection_pool import connection_pool
    connector = connection_pool.get(cid)
    inserted = pool_loop.run_until_complete(connector.batch_insert(database, table_name, rows))
    console.print(f"[green]✓ Inserted {inserted} rows into {table}[/green]")


@data_app.command("compare")
def data_compare(
    source: str = typer.Argument(..., help="Source table (format: conn:db.table)"),
    target: str = typer.Argument(..., help="Target table (format: conn:db.table)"),
    output: str = typer.Option("", "--output", "-o", help="Save sync SQL to file"),
) -> None:
    """Compare row-level data between two tables (same or different connections)."""
    from open_navicat.services.data_sync_engine import data_sync_engine

    def parse_table(spec: str) -> tuple[str, str, str]:
        parts = spec.split(":", 1)
        if len(parts) == 1:
            conn_id = _get_active_conn()
            db, tbl = parts[0].split(".", 1)
        else:
            conn_id = _resolve_conn(parts[0])
            db, tbl = parts[1].split(".", 1)
        return conn_id, db, tbl

    src_conn, src_db, src_tbl = parse_table(source)
    tgt_conn, tgt_db, tgt_tbl = parse_table(target)

    console.print(f"[yellow]Comparing {src_db}.{src_tbl} → {tgt_db}.{tgt_tbl}...[/yellow]")
    result = data_sync_engine.compare_tables(src_conn, src_db, src_tbl, tgt_conn, tgt_db, tgt_tbl)
    console.print(f"[bold]Source:[/bold] {result.source_rows} rows, [bold]Target:[/bold] {result.target_rows} rows")
    console.print(f"[bold]Differences:[/bold] {result.total_diffs} total")
    console.print(f"  [green]+{len(result.inserts)}[/green] inserts  "
                  f"[yellow]~{len(result.updates)}[/yellow] updates  "
                  f"[red]-{len(result.deletes)}[/red] deletes")
    if result.total_diffs == 0:
        console.print("[green]✓ Tables are in sync![/green]")
        return

    script = data_sync_engine.generate_sync_script(result)
    if output:
        with open(output, "w", encoding="utf-8") as f:
            f.write(script)
        console.print(f"[green]Sync script saved to {output}[/green]")
    else:
        from rich.syntax import Syntax
        console.print(Syntax(script, "sql", theme="monokai"))


@data_app.command("sync")
def data_sync(
    source: str = typer.Argument(..., help="Source table (format: conn:db.table)"),
    target: str = typer.Argument(..., help="Target table (format: conn:db.table)"),
    dry_run: bool = typer.Option(True, "--dry-run/--execute", help="Preview only (default: dry-run)"),
) -> None:
    """Execute data sync: apply INSERT/UPDATE/DELETE to make target match source."""
    from open_navicat.services.data_sync_engine import data_sync_engine

    def parse_table(spec: str) -> tuple[str, str, str]:
        parts = spec.split(":", 1)
        if len(parts) == 1:
            conn_id = _get_active_conn()
            db, tbl = parts[0].split(".", 1)
        else:
            conn_id = _resolve_conn(parts[0])
            db, tbl = parts[1].split(".", 1)
        return conn_id, db, tbl

    src_conn, src_db, src_tbl = parse_table(source)
    tgt_conn, tgt_db, tgt_tbl = parse_table(target)

    console.print(f"[yellow]Comparing {src_db}.{src_tbl} → {tgt_db}.{tgt_tbl}...[/yellow]")
    result = data_sync_engine.compare_tables(src_conn, src_db, src_tbl, tgt_conn, tgt_db, tgt_tbl)

    if result.total_diffs == 0:
        console.print("[green]✓ Tables are already in sync![/green]")
        return

    console.print(f"Changes to apply: [green]+{len(result.inserts)}[/green] "
                  f"[yellow]~{len(result.updates)}[/yellow] "
                  f"[red]-{len(result.deletes)}[/red]")

    if dry_run:
        script = data_sync_engine.generate_sync_script(result)
        from rich.syntax import Syntax
        console.print(Syntax(script, "sql", theme="monokai"))
        console.print("[yellow]Dry-run mode. Use --execute to apply changes.[/yellow]")
        return

    if not typer.confirm("Apply these changes to target?"):
        console.print("[yellow]Cancelled.[/yellow]")
        return

    from open_navicat.dal.connection_pool import _loop as pool_loop
    from open_navicat.dal.connection_pool import connection_pool
    tgt_connector = connection_pool.get(tgt_conn)
    if not tgt_connector:
        console.print("[red]Target connection not found.[/red]")
        raise typer.Exit(1)

    script = data_sync_engine.generate_sync_script(result)
    for stmt in script.split("\n"):
        stmt = stmt.strip()
        if stmt:
            pool_loop.run_until_complete(tgt_connector.execute(stmt))
    console.print(f"[green]✓ Synced {result.total_diffs} changes to target.[/green]")
