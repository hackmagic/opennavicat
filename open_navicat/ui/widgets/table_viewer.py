"""Data viewer — grid-style table browser with pagination, sorting, and inline editing."""

from __future__ import annotations

import logging

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QDialog,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QToolBar,
    QVBoxLayout,
    QWidget,
)

from open_navicat.dal.connection_pool import _loop as pool_loop
from open_navicat.dal.connection_pool import connection_pool
from open_navicat.i18n import t
from open_navicat.models import QueryResult

_log = logging.getLogger(__name__)


def _fmt_bytes(n: int | float) -> str:
    """Format byte count to human-readable string."""
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if abs(n) < 1024:
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} PB"


class TableViewerWidget(QWidget):
    """Displays table data in a spreadsheet-like grid with filtering and pagination."""

    PAGE_SIZES = [100, 200, 500, 1000, 5000]

    def __init__(
        self,
        connection_id: str,
        database: str,
        table: str,
        parent=None,
    ) -> None:
        super().__init__(parent)
        self._connection_id = connection_id
        self._database = database
        self._table = table

        from open_navicat.config import config as _cfg
        self._page_size = _cfg.get("records.limit", 500) if _cfg.get("records.limit_enabled", True) else 500
        self._current_page = 0
        self._total_rows = 0
        self._filters: dict[str, str] = {}
        self._order_by: str | None = None
        self._order_dir: str = "ASC"
        self._loaded_column_names: list[str] = []

        self._setup_ui()
        self._load_page()

    def _setup_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(4, 4, 4, 4)

        # -- Toolbar --
        toolbar = QToolBar(self)

        # Group 1: Profile & Transaction
        self._btn_profile = QPushButton(t("data_viewer.table_profile"), self)
        self._btn_profile.clicked.connect(self._show_table_profile)
        toolbar.addWidget(self._btn_profile)

        self._btn_commit = QPushButton(t("data_viewer.commit"), self)
        self._btn_commit.clicked.connect(self._commit_transaction)
        toolbar.addWidget(self._btn_commit)

        self._btn_rollback = QPushButton(t("data_viewer.rollback"), self)
        self._btn_rollback.clicked.connect(self._rollback_transaction)
        toolbar.addWidget(self._btn_rollback)

        toolbar.addSeparator()

        # Group 2: Edit operations
        self._btn_refresh = QPushButton(t("data_viewer.refresh"), self)
        self._btn_refresh.clicked.connect(self._load_page)
        toolbar.addWidget(self._btn_refresh)

        self._btn_add = QPushButton(t("data_viewer.add_row"), self)
        self._btn_add.clicked.connect(self._add_row)
        toolbar.addWidget(self._btn_add)

        self._btn_delete = QPushButton(t("data_viewer.delete"), self)
        self._btn_delete.clicked.connect(self._delete_selected)
        toolbar.addWidget(self._btn_delete)

        self._btn_cell_editor = QPushButton(t("data_viewer.cell_editor"), self)
        self._btn_cell_editor.setCheckable(True)
        self._btn_cell_editor.setChecked(True)
        self._btn_cell_editor.clicked.connect(self._toggle_cell_editor)
        toolbar.addWidget(self._btn_cell_editor)

        toolbar.addSeparator()

        # Group 3: Filter, Sort, Columns
        toolbar.addWidget(QLabel(t("data_viewer.filter") + ":", self))
        self._filter_input = QLineEdit(self)
        self._filter_input.setPlaceholderText("column = value  (SQL WHERE clause)")
        self._filter_input.setMinimumWidth(260)
        self._filter_input.returnPressed.connect(self._apply_filter)
        toolbar.addWidget(self._filter_input)

        self._btn_filter = QPushButton(t("data_viewer.apply"), self)
        self._btn_filter.clicked.connect(self._apply_filter)
        toolbar.addWidget(self._btn_filter)

        self._btn_sort = QPushButton(t("data_viewer.sort"), self)
        self._btn_sort.setCheckable(True)
        self._btn_sort.clicked.connect(self._show_sort_popup)
        toolbar.addWidget(self._btn_sort)

        self._btn_columns = QPushButton(t("data_viewer.columns"), self)
        self._btn_columns.clicked.connect(self._show_columns_manager)
        toolbar.addWidget(self._btn_columns)

        toolbar.addSeparator()

        # Group 4: Import, Export, Analysis, Gen, BI
        self._btn_export = QPushButton("📤 " + t("menu.file.export"), self)
        self._btn_export.clicked.connect(self._export_data)
        toolbar.addWidget(self._btn_export)

        self._btn_import = QPushButton("📥 " + t("menu.file.import"), self)
        self._btn_import.clicked.connect(self._import_data)
        toolbar.addWidget(self._btn_import)

        self._btn_gen_data = QPushButton(t("data_viewer.gen_data"), self)
        self._btn_gen_data.clicked.connect(self._generate_test_data)
        toolbar.addWidget(self._btn_gen_data)

        self._btn_analysis = QPushButton(t("data_viewer.data_analysis"), self)
        self._btn_analysis.clicked.connect(self._show_data_analysis)
        toolbar.addWidget(self._btn_analysis)

        self._btn_bi = QPushButton(t("data_viewer.bi_workspace"), self)
        self._btn_bi.clicked.connect(self._show_bi_workspace)
        toolbar.addWidget(self._btn_bi)

        toolbar.addSeparator()
        toolbar.addWidget(QLabel(t("data_viewer.per_page") + ":", self))
        self._combo_page_size = QComboBox(self)
        self._combo_page_size.addItems([str(s) for s in self.PAGE_SIZES])
        self._combo_page_size.setCurrentText(str(self._page_size))
        self._combo_page_size.currentTextChanged.connect(self._change_page_size)
        toolbar.addWidget(self._combo_page_size)

        toolbar.addSeparator()
        self._status_msg = QLabel("", self)
        self._status_msg.setStyleSheet("color: #888; font-size: 11px;")
        toolbar.addWidget(self._status_msg)

        layout.addWidget(toolbar)

        # -- Table grid --
        self._table_widget = QTableWidget(self)
        self._table_widget.setAlternatingRowColors(True)
        self._table_widget.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows
        )
        self._table_widget.setSelectionMode(
            QAbstractItemView.SelectionMode.ExtendedSelection
        )
        self._table_widget.horizontalHeader().setStretchLastSection(True)
        self._table_widget.horizontalHeader().setSectionsMovable(True)
        self._table_widget.setSortingEnabled(True)
        self._table_widget.horizontalHeader().sectionClicked.connect(self._on_header_clicked)
        self._table_widget.cellChanged.connect(self._on_cell_changed)
        layout.addWidget(self._table_widget)

        # -- Pagination bar --
        pagination = QHBoxLayout()
        self._btn_first = QPushButton("⏮ First", self)
        self._btn_first.clicked.connect(lambda: self._go_to_page(0))
        pagination.addWidget(self._btn_first)

        self._btn_prev = QPushButton("◀ Prev", self)
        self._btn_prev.clicked.connect(self._prev_page)
        pagination.addWidget(self._btn_prev)

        self._page_info = QLabel("Page 0 of 0", self)
        pagination.addWidget(self._page_info)

        self._btn_next = QPushButton("Next ▶", self)
        self._btn_next.clicked.connect(self._next_page)
        pagination.addWidget(self._btn_next)

        self._btn_last = QPushButton("Last ⏭", self)
        self._btn_last.clicked.connect(self._go_to_last_page)
        pagination.addWidget(self._btn_last)

        pagination.addStretch()
        self._row_count_label = QLabel("", self)
        pagination.addWidget(self._row_count_label)

        layout.addLayout(pagination)

    # ---- data loading ----

    def _on_header_clicked(self, col: int) -> None:
        """Toggle sort by column."""
        header = self._table_widget.horizontalHeaderItem(col)
        if not header:
            return
        col_name = header.text()
        if self._order_by == col_name:
            self._order_dir = "DESC" if self._order_dir == "ASC" else "ASC"
        else:
            self._order_by = col_name
            self._order_dir = "ASC"
        self._current_page = 0
        self._load_page()

    def _load_page(self) -> None:
        connector = connection_pool.get(self._connection_id)
        if not connector:
            return

        # Build WHERE clause from filters
        where_clause = ""
        if self._filters:
            where_parts = []
            for k, v in self._filters.items():
                # v can be: plain value (=), or "LIKE x", ">= 5", "!= x" etc.
                if any(v.startswith(op) for op in [">=", "<=", "!=", ">", "<", "LIKE", "IN", "BETWEEN"]):
                    where_parts.append(f"`{k}` {v}")
                else:
                    where_parts.append(f"`{k}` = '{v}'")
            where_clause = "WHERE " + " AND ".join(where_parts)

        try:
            # Count total first
            count_sql = f"SELECT COUNT(*) FROM `{self._database}`.`{self._table}` {where_clause}"
            count_result = pool_loop.run_until_complete(connector.execute(count_sql))
            self._total_rows = count_result.rows[0][0] if count_result.rows else 0

            # Fetch page
            offset = self._current_page * self._page_size
            order = f"ORDER BY `{self._order_by}` {self._order_dir}" if self._order_by else ""
            page_sql = (
                f"SELECT * FROM `{self._database}`.`{self._table}` {where_clause} "
                f"{order} LIMIT {self._page_size} OFFSET {offset}"
            )
            result = pool_loop.run_until_complete(connector.execute(page_sql))

            self._render(result)
            self._update_pagination()
        except Exception as e:
            self._table_widget.setRowCount(0)
            self._table_widget.setColumnCount(0)
            self._row_count_label.setText(f"Error: {e}")

    def _render(self, result: QueryResult) -> None:
        self._table_widget.blockSignals(True)

        from open_navicat.config import config as _cfg
        thousand_sep = _cfg.get("records.thousand_sep", False)
        null_str = _cfg.get("data_viewer.null_string", "(NULL)")


        if result.is_select:
            cols = [c.name for c in result.columns]
            self._loaded_column_names = cols
            self._table_widget.setColumnCount(len(cols))
            self._table_widget.setHorizontalHeaderLabels(cols)
            self._table_widget.setRowCount(len(result.rows))

            for row_idx, row in enumerate(result.rows):
                for col_idx, val in enumerate(row):
                    if val is None:
                        display = null_str or "(NULL)"
                    elif isinstance(val, (int, float)) and thousand_sep:
                        display = f"{val:,}"
                    else:
                        display = str(val)
                    item = QTableWidgetItem(display)
                    if val is None:
                        item.setForeground(QColor("#808080"))
                    self._table_widget.setItem(row_idx, col_idx, item)

            self._table_widget.horizontalHeader().resizeSections()
        else:
            self._table_widget.setRowCount(0)
            self._table_widget.setColumnCount(0)

        self._table_widget.blockSignals(False)

    def _update_pagination(self) -> None:
        total_pages = max(1, (self._total_rows + self._page_size - 1) // self._page_size)
        self._page_info.setText(
            f"Page {self._current_page + 1} of {total_pages}"
        )
        self._row_count_label.setText(f"Total rows: {self._total_rows}")

        self._btn_first.setEnabled(self._current_page > 0)
        self._btn_prev.setEnabled(self._current_page > 0)
        self._btn_next.setEnabled(self._current_page < total_pages - 1)
        self._btn_last.setEnabled(self._current_page < total_pages - 1)

    # ---- pagination actions ----

    def _go_to_page(self, page: int) -> None:
        self._current_page = page
        self._load_page()

    def _prev_page(self) -> None:
        if self._current_page > 0:
            self._current_page -= 1
            self._load_page()

    def _next_page(self) -> None:
        total_pages = max(1, (self._total_rows + self._page_size - 1) // self._page_size)
        if self._current_page < total_pages - 1:
            self._current_page += 1
            self._load_page()

    def _go_to_last_page(self) -> None:
        total_pages = max(1, (self._total_rows + self._page_size - 1) // self._page_size)
        self._current_page = total_pages - 1
        self._load_page()

    def _change_page_size(self, text: str) -> None:
        try:
            self._page_size = int(text)
            self._current_page = 0
            self._load_page()
        except ValueError:
            pass

    # ---- filter ----

    def _apply_filter(self) -> None:
        filter_text = self._filter_input.text().strip()
        if filter_text:
            # Support operators: =, !=, >, <, >=, <=, LIKE, IN, BETWEEN
            for op in ["!=", ">=", "<=", "=", ">", "<"]:
                if op in filter_text:
                    parts = filter_text.split(op, 1)
                    col = parts[0].strip().strip("`\"'")
                    val = parts[1].strip().strip("'\"")
                    if op == "=":
                        self._filters[col] = val
                    else:
                        self._filters[col] = f"{op} {val}"
                    break
            else:
                # Check for LIKE / IN / BETWEEN
                import re
                m = re.match(r"(\w+)\s+(LIKE|IN|BETWEEN)\s+(.+)", filter_text, re.IGNORECASE)
                if m:
                    self._filters[m.group(1)] = f"{m.group(2)} {m.group(3).strip()}"
                else:
                    self._filters.clear()
        else:
            self._filters.clear()
        self._current_page = 0
        self._load_page()

    # ---- CRUD operations ----

    def _on_cell_changed(self, row: int, col: int) -> None:
        """Update cell value in database when user finishes editing."""
        connector = connection_pool.get(self._connection_id)
        if not connector:
            return

        item = self._table_widget.item(row, col)
        if not item:
            return

        new_value = item.text()
        # Get column name
        col_name = self._table_widget.horizontalHeaderItem(col).text()
        if not col_name:
            return

        # Get primary key column(s) to identify the row
        pk_col = self._table_widget.horizontalHeaderItem(0).text()

        # Get the PK value for this row (first column)
        pk_item = self._table_widget.item(row, 0)
        if not pk_item:
            return
        pk_value = pk_item.text()

        sql = f"UPDATE `{self._database}`.`{self._table}` SET `{col_name}` = %s WHERE `{pk_col}` = %s"
        try:
            pool_loop.run_until_complete(connector.execute(sql, (new_value, pk_value)))
        except Exception as e:
            item.setToolTip(f"Update failed: {e}")

    def _add_row(self) -> None:
        """Insert a new blank row into the table, then reload."""
        connector = connection_pool.get(self._connection_id)
        if not connector:
            return

        # Build INSERT with default values — insert one row with no columns specified
        sql = f"INSERT INTO `{self._database}`.`{self._table}` () VALUES ()"
        try:
            result = pool_loop.run_until_complete(connector.execute(sql))
            if result.insert_id:
                self._status_msg.setText(f"Inserted row id={result.insert_id}")
            # Reload to show the new row
            self._current_page = 0
            self._load_page()
        except Exception:
            # If bare INSERT fails (e.g. NOT NULL columns), use a smarter approach:
            # get column names and insert with NULL/defaults for each
            try:
                info_sql = (
                    "SELECT COLUMN_NAME FROM information_schema.COLUMNS "
                    "WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s "
                    "ORDER BY ORDINAL_POSITION"
                )
                rows = pool_loop.run_until_complete(
                    connector._fetch_all(info_sql, (self._database, self._table))
                )
                cols = [r[0] for r in rows]
                placeholders = ", ".join(["NULL"] * len(cols))
                col_names = ", ".join(f"`{c}`" for c in cols)
                sql2 = f"INSERT INTO `{self._database}`.`{self._table}` ({col_names}) VALUES ({placeholders})"
                pool_loop.run_until_complete(connector.execute(sql2))
                self._current_page = 0
                self._load_page()
            except Exception as e2:
                self._status_msg.setText(f"Add row failed: {e2}")

    def _delete_selected(self) -> None:
        """Delete selected rows from the table."""
        selected = self._table_widget.selectedItems()
        if not selected:
            return

        rows_to_delete = set()
        for item in selected:
            rows_to_delete.add(item.row())

        if not rows_to_delete:
            return

        from PySide6.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            self, t("browser.truncate_table"),
            f"确定要删除选中的 {len(rows_to_delete)} 行吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        connector = connection_pool.get(self._connection_id)
        if not connector:
            return

        pk_col = self._table_widget.horizontalHeaderItem(0).text()
        deleted = 0
        for row_idx in sorted(rows_to_delete, reverse=True):
            pk_item = self._table_widget.item(row_idx, 0)
            if not pk_item:
                continue
            pk_val = pk_item.text()
            sql = f"DELETE FROM `{self._database}`.`{self._table}` WHERE `{pk_col}` = %s"
            try:
                result = pool_loop.run_until_complete(connector.execute(sql, (pk_val,)))
                deleted += result.affected_rows
            except Exception as e:
                self._status_msg.setText(f"Delete failed: {e}")

        self._status_msg.setText(f"Deleted {deleted} row(s)")
        self._load_page()

    # ---- export / import ----

    def _export_data(self) -> None:
        """Export current page or all data to CSV or Excel."""
        from PySide6.QtWidgets import QFileDialog, QMessageBox
        path, _ = QFileDialog.getSaveFileName(
            self, "导出数据", f"{self._table}",
            "CSV 文件 (*.csv);;Excel 文件 (*.xlsx);;所有文件 (*)",
        )
        if not path:
            return

        connector = connection_pool.get(self._connection_id)
        if not connector:
            return

        try:
            sql = f"SELECT * FROM `{self._database}`.`{self._table}`"
            where = ""
            if self._filters:
                where = " WHERE " + " AND ".join(f"{k} = '{v}'" for k, v in self._filters.items() if not any(v.startswith(op) for op in [">=", "<=", "!=", ">", "<", "LIKE", "IN"]))
                for k, v in self._filters.items():
                    if any(v.startswith(op) for op in [">=", "<=", "!=", ">", "<", "LIKE", "IN"]):
                        where = f" WHERE `{k}` {v}"
            sql += where
            result = pool_loop.run_until_complete(connector.execute(sql))

            if not result.columns:
                return

            if path.endswith(".xlsx"):
                self._export_excel(path, result)
            else:
                self._export_csv(path, result)

            self._status_msg.setText(f"✅ 已导出 {len(result.rows)} 行到 {path}")
        except Exception as e:
            QMessageBox.warning(self, "导出失败", str(e))

    def _export_csv(self, path: str, result) -> None:
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([c.name for c in result.columns])
            for row in result.rows:
                writer.writerow([str(v) if v is not None else "" for v in row])

    def _export_excel(self, path: str, result) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = self._table
        ws.append([c.name for c in result.columns])
        for row in result.rows:
            ws.append([v for v in row])
        wb.save(path)

    def _import_data(self) -> None:
        """Import data from CSV or Excel file into the table."""
        from PySide6.QtWidgets import QFileDialog, QMessageBox
        path, _ = QFileDialog.getOpenFileName(
            self, "导入数据", "",
            "CSV 文件 (*.csv);;Excel 文件 (*.xlsx);;所有文件 (*)",
        )
        if not path:
            return

        try:
            rows = []
            if path.endswith(".xlsx"):
                from openpyxl import load_workbook
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [str(v) if v is not None else None for v in row])))
            else:
                import csv
                with open(path, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    rows = [dict(r) for r in reader]

            if not rows:
                QMessageBox.information(self, "导入", "文件为空。")
                return

            connector = connection_pool.get(self._connection_id)
            if not connector:
                return

            batch_size = 100
            total = 0
            for i in range(0, len(rows), batch_size):
                batch = rows[i:i + batch_size]
                inserted = pool_loop.run_until_complete(
                    connector.batch_insert(self._database, self._table, batch)
                )
                total += inserted

            self._status_msg.setText(f"✅ 已导入 {total} 行")
            self._load_page()
        except Exception as e:
            QMessageBox.warning(self, "导入失败", str(e))

    def _generate_test_data(self) -> None:
        """Generate test data using AI and insert into the table."""
        from PySide6.QtWidgets import QInputDialog

        from open_navicat.dal.connection_pool import _loop as pool_loop
        from open_navicat.dal.connection_pool import connection_pool
        from open_navicat.services.ai_service import ai_service

        count, ok = QInputDialog.getInt(self, "生成测试数据", "生成行数:", 10, 1, 1000)
        if not ok:
            return

        connector = connection_pool.get(self._connection_id)
        if not connector:
            return

        self._status_msg.setText("⏳ 正在生成数据...")
        try:
            # Get table info for schema
            info_sql = "SELECT COLUMN_NAME, COLUMN_TYPE, IS_NULLABLE, COLUMN_DEFAULT, EXTRA, COLUMN_COMMENT FROM information_schema.COLUMNS WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s ORDER BY ORDINAL_POSITION"
            rows = pool_loop.run_until_complete(connector._fetch_all(info_sql, (self._database, self._table)))
            from open_navicat.models.table_schema import ColumnInfo, TableInfo
            table_info = TableInfo(name=self._table, database=self._database)
            for r in rows:
                table_info.columns.append(ColumnInfo(
                    name=r[0], data_type=r[1], nullable=(r[2] == "YES"),
                    default=r[3], is_auto_increment="auto_increment" in (r[4] or ""),
                    comment=r[5] or "",
                ))

            # Generate data via AI
            generated = ai_service.generate_data(table_info, count)
            if not generated:
                self._status_msg.setText("⚠️ 数据生成失败")
                return

            # Insert
            total = pool_loop.run_until_complete(
                connector.batch_insert(self._database, self._table, generated)
            )
            self._status_msg.setText(f"✅ 已生成并插入 {total} 行测试数据")
            self._load_page()
        except Exception as e:
            self._status_msg.setText(f"❌ 生成失败: {e}")

    # ---- new toolbar actions ----

    def _show_table_profile(self) -> None:
        """Show table metadata: engine, row count, data size, indexes, etc."""
        connector = connection_pool.get(self._connection_id)
        if not connector:
            return
        try:
            sql = (
                "SELECT TABLE_NAME, ENGINE, TABLE_ROWS, DATA_LENGTH, INDEX_LENGTH, "
                "CREATE_TIME, UPDATE_TIME, TABLE_COLLATION, TABLE_COMMENT "
                "FROM information_schema.TABLES "
                "WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s"
            )
            r = pool_loop.run_until_complete(connector.execute(sql, (self._database, self._table)))
            if not r.rows:
                QMessageBox.information(self, "表信息", "无法获取表信息。")
                return
            row = r.rows[0]
            dlg = QDialog(self)
            dlg.setWindowTitle(f"表信息 — {self._table}")
            dlg.resize(450, 300)
            layout = QVBoxLayout(dlg)
            items = [
                ("表名", row[0]),
                ("引擎", row[1] or "—"),
                ("行数(估算)", f"{row[2]:,}" if row[2] is not None else "—"),
                ("数据大小", _fmt_bytes(row[3]) if row[3] else "—"),
                ("索引大小", _fmt_bytes(row[4]) if row[4] else "—"),
                ("创建时间", str(row[5] or "—")),
                ("更新时间", str(row[6] or "—")),
                ("排序规则", row[7] or "—"),
                ("注释", row[8] or "—"),
            ]
            for label, val in items:
                row_w = QHBoxLayout()
                lbl = QLabel(f"{label}:", dlg)
                lbl.setFixedWidth(100)
                row_w.addWidget(lbl)
                val_lbl = QLabel(str(val), dlg)
                val_lbl.setWordWrap(True)
                row_w.addWidget(val_lbl, 1)
                layout.addLayout(row_w)
            layout.addStretch()
            btn = QPushButton("关闭", dlg)
            btn.clicked.connect(dlg.accept)
            layout.addWidget(btn, 0, Qt.AlignmentFlag.AlignRight)
            dlg.exec()
        except Exception as e:
            QMessageBox.warning(self, "错误", str(e))

    def _commit_transaction(self) -> None:
        connector = connection_pool.get(self._connection_id)
        if connector:
            try:
                pool_loop.run_until_complete(connector.execute("COMMIT"))
                self._status_msg.setText("✅ 已提交")
            except Exception as e:
                self._status_msg.setText(f"❌ 提交失败: {e}")

    def _rollback_transaction(self) -> None:
        connector = connection_pool.get(self._connection_id)
        if connector:
            try:
                pool_loop.run_until_complete(connector.execute("ROLLBACK"))
                self._status_msg.setText("✅ 已回滚")
            except Exception as e:
                self._status_msg.setText(f"❌ 回滚失败: {e}")

    def _toggle_cell_editor(self, checked: bool) -> None:
        if checked:
            self._table_widget.setEditTriggers(
                QAbstractItemView.EditTrigger.DoubleClicked
                | QAbstractItemView.EditTrigger.EditKeyPressed
            )
            self._status_msg.setText("单元格编辑: 已启用")
        else:
            self._table_widget.setEditTriggers(
                QAbstractItemView.EditTrigger.NoEditTriggers
            )
            self._status_msg.setText("单元格编辑: 已禁用")

    def _show_sort_popup(self) -> None:
        from PySide6.QtWidgets import QMenu
        menu = QMenu(self)
        for i in range(self._table_widget.columnCount()):
            h = self._table_widget.horizontalHeaderItem(i)
            if not h:
                continue
            col_name = h.text()
            asc_act = menu.addAction(f"↑ {col_name} ASC")
            asc_act.triggered.connect(lambda checked, c=col_name: self._apply_sort(c, "ASC"))
            desc_act = menu.addAction(f"↓ {col_name} DESC")
            desc_act.triggered.connect(lambda checked, c=col_name: self._apply_sort(c, "DESC"))
        btn = self.sender()
        if btn:
            menu.exec(btn.mapToGlobal(btn.rect().bottomLeft()))

    def _apply_sort(self, col_name: str, direction: str) -> None:
        self._order_by = col_name
        self._order_dir = direction
        self._current_page = 0
        self._load_page()

    def _show_columns_manager(self) -> None:
        from PySide6.QtWidgets import QMenu
        menu = QMenu(self)
        actions = []
        for i in range(self._table_widget.columnCount()):
            h = self._table_widget.horizontalHeaderItem(i)
            if not h:
                continue
            act = menu.addAction(h.text())
            act.setCheckable(True)
            act.setChecked(not self._table_widget.isColumnHidden(i))
            actions.append((i, act))
        for idx, act in actions:
            act.triggered.connect(lambda checked, col=idx: self._table_widget.setColumnHidden(col, not checked))
        btn = self.sender()
        if btn:
            menu.exec(btn.mapToGlobal(btn.rect().bottomLeft()))

    def _show_data_analysis(self) -> None:
        """Run analytical queries against the table and show results in a dialog."""
        connector = connection_pool.get(self._connection_id)
        if not connector:
            return
        dlg = QDialog(self)
        dlg.setWindowTitle(f"数据分析 — {self._table}")
        dlg.resize(780, 520)
        layout = QVBoxLayout(dlg)
        tabs = QTabWidget(dlg)

        try:
            db_table = f"`{self._database}`.`{self._table}`"

            # ── Tab 1: Overview ──
            overview = QTableWidget(tabs)
            overview.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            tabs.addTab(overview, "概览")

            stats_sql = f"SELECT COUNT(*) as total FROM {db_table}"
            pool_loop.run_until_complete(connector.execute(stats_sql))

            # Per-column stats via information_schema
            info_sql = ("SELECT COLUMN_NAME, DATA_TYPE, IS_NULLABLE, "
                        "CHARACTER_MAXIMUM_LENGTH, NUMERIC_PRECISION, NUMERIC_SCALE "
                        "FROM information_schema.COLUMNS "
                        "WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s "
                        "ORDER BY ORDINAL_POSITION")
            col_meta = pool_loop.run_until_complete(
                connector.execute(info_sql, (self._database, self._table)))
            col_types = {}
            for r in (col_meta.rows or []):
                col_types[r[0]] = r[1].lower() if r[1] else "text"

            # Build column stats
            overview.setColumnCount(6)
            overview.setHorizontalHeaderLabels(
                ["列名", "类型", "行数", "非空", "唯一值", "空值"])
            overview.setRowCount(len(self._loaded_column_names))

            numeric_types = {"int", "tinyint", "smallint", "mediumint", "bigint",
                             "float", "double", "decimal", "real"}
            data = []
            for i, col_name in enumerate(self._loaded_column_names):
                dtype = col_types.get(col_name, "text")
                try:
                    q = f"SELECT COUNT(*), COUNT(`{col_name}`), COUNT(DISTINCT `{col_name}`) FROM {db_table}"
                    r = pool_loop.run_until_complete(connector.execute(q))
                    if r.rows:
                        cnt, nonnull, distinct = r.rows[0]
                        null_cnt = (cnt or 0) - (nonnull or 0)
                        overview.setItem(i, 0, QTableWidgetItem(col_name))
                        overview.setItem(i, 1, QTableWidgetItem(dtype))
                        overview.setItem(i, 2, QTableWidgetItem(str(cnt or 0)))
                        overview.setItem(i, 3, QTableWidgetItem(str(nonnull or 0)))
                        overview.setItem(i, 4, QTableWidgetItem(str(distinct or 0)))
                        overview.setItem(i, 5, QTableWidgetItem(str(null_cnt)))
                        data.append((col_name, dtype, cnt, nonnull, distinct, null_cnt))

                        # Numeric stats
                        if dtype in numeric_types and nonnull and nonnull > 0:
                            q2 = f"SELECT MIN(`{col_name}`), MAX(`{col_name}`), AVG(`{col_name}`) FROM {db_table}"
                            r2 = pool_loop.run_until_complete(connector.execute(q2))
                            if r2.rows and r2.rows[0][0] is not None:
                                overview.setColumnCount(9)
                                overview.setHorizontalHeaderLabels(
                                    ["列名", "类型", "行数", "非空", "唯一值", "空值", "最小值", "最大值", "平均值"])
                                mn, mx, avg = r2.rows[0]
                                overview.setItem(i, 6, QTableWidgetItem(str(mn)))
                                overview.setItem(i, 7, QTableWidgetItem(str(mx)))
                                overview.setItem(i, 8, QTableWidgetItem(f"{avg:.4f}" if isinstance(avg, float) else str(avg)))
                except Exception as e:
                    _log.debug("Column stats query failed for %s: %s", col_name, e)
            overview.resizeColumnsToContents()

            # ── Tab 2: Value Distribution (top 10 per column) ──
            freq_tab = QTableWidget(tabs)
            freq_tab.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            freq_tab.setStyleSheet(overview.styleSheet())
            tabs.addTab(freq_tab, "值分布")

            freq_parts: list[list[str]] = []
            for col_name in self._loaded_column_names:
                try:
                    q = (f"SELECT `{col_name}`, COUNT(*) as cnt FROM {db_table} "
                         f"WHERE `{col_name}` IS NOT NULL "
                         f"GROUP BY `{col_name}` ORDER BY cnt DESC LIMIT 10")
                    r = pool_loop.run_until_complete(connector.execute(q))
                    if r.rows:
                        freq_parts.append(f"── {col_name} ──")
                        for row in r.rows:
                            freq_parts.append(f"  {row[0]}: {row[1]}")
                        freq_parts.append("")
                except Exception as e:
                    _log.debug("Frequency query failed for %s: %s", col_name, e)

            freq_tab.setColumnCount(1)
            freq_tab.setHorizontalHeaderLabels(["值 (频次)"])
            freq_tab.setRowCount(len(freq_parts))
            for i, line in enumerate(freq_parts):
                item = QTableWidgetItem(line)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                freq_tab.setItem(i, 0, item)
            freq_tab.resizeColumnsToContents()

        except Exception as e:
            error_widget = QLabel(f"分析出错: {e}", tabs)
            error_widget.setStyleSheet("color: #f44747; padding: 20px;")
            tabs.addTab(error_widget, "错误")

        layout.addWidget(tabs, 1)
        btn = QPushButton("关闭", dlg)
        btn.setStyleSheet(
            "background: #3c3c3c; color: #ccc; border: 1px solid #555; "
            "border-radius: 3px; padding: 4px 16px;")
        btn.clicked.connect(dlg.accept)
        layout.addWidget(btn, 0, Qt.AlignmentFlag.AlignRight)
        dlg.exec()

    def _show_bi_workspace(self) -> None:
        mw = self.window()
        if hasattr(mw, '_show_bi_dashboard'):
            mw._show_bi_dashboard()
        else:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.information(self, t("data_viewer.bi_workspace"), "敬请期待")
