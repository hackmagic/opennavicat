"""Import/Export Wizard — multi-step dialogs for data import and export."""

from __future__ import annotations

import csv

from PySide6.QtWidgets import (
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QRadioButton,
    QSpinBox,
    QVBoxLayout,
    QWizard,
    QWizardPage,
)

from open_navicat.dal.connection_pool import _loop as pool_loop
from open_navicat.dal.connection_pool import connection_pool
from open_navicat.i18n import t


class FormatPage(QWizardPage):
    """Page 1: select import/export format."""

    FORMATS = [
        ("txt", t("import_export.file_type.text")),
        ("csv", t("import_export.file_type.csv")),
        ("json", t("import_export.file_type.json")),
        ("xml", t("import_export.file_type.xml")),
        ("html", t("import_export.file_type.html")),
        ("sql", t("import_export.file_type.sql")),
        ("xlsx", t("import_export.file_type.excel")),
    ]

    def __init__(self, is_import: bool = True) -> None:
        super().__init__()
        self.setTitle(t("import_export.step_format"))
        self._group = QButtonGroup(self)
        layout = QVBoxLayout(self)
        for i, (_, label) in enumerate(self.FORMATS):
            rb = QRadioButton(label)
            layout.addWidget(rb)
            self._group.addButton(rb, i)
        self._group.buttons()[1].setChecked(True)  # Default: CSV

    def selected_format(self) -> str:
        return self.FORMATS[self._group.checkedId()][0]


class FilePage(QWizardPage):
    """Page 2: select source/target file path."""

    def __init__(self, is_import: bool) -> None:
        super().__init__()
        self._is_import = is_import
        self.setTitle(t("import_export.step_file"))
        layout = QFormLayout(self)
        self._path_input = QLineEdit(self)
        self._path_input.setPlaceholderText(t("import_export.file_placeholder"))
        browse_btn = QPushButton(t("import_export.btn.browse"), self)
        browse_btn.clicked.connect(self._browse)
        row = QHBoxLayout()
        row.addWidget(self._path_input)
        row.addWidget(browse_btn)
        layout.addRow(t("import_export.label.file"), row)
        self.registerField("file_path*", self._path_input)

    def _browse(self) -> None:
        filter_str = "文本文件 (*.txt);;CSV (*.csv);;JSON (*.json);;XML (*.xml);;HTML (*.html);;SQL (*.sql);;Excel (*.xlsx);;所有文件 (*)"
        if self._is_import:
            path, _ = QFileDialog.getOpenFileName(self, t("import_export.import_file_dialog"), "", filter_str)
        else:
            path, _ = QFileDialog.getSaveFileName(self, t("import_export.export_file_dialog"), "export", filter_str)
        if path:
            self._path_input.setText(path)

    def file_path(self) -> str:
        return self._path_input.text()


class OptionsPage(QWizardPage):
    """Page 3: encoding, delimiter, header row, etc."""

    def __init__(self) -> None:
        super().__init__()
        self.setTitle(t("import_export.step_options"))
        layout = QFormLayout(self)

        self._encoding_combo = QComboBox(self)
        self._encoding_combo.addItems(["utf-8", "utf-8-sig", "gbk", "gb2312", "latin-1"])
        self._encoding_combo.setCurrentText("utf-8-sig" if "import" in __name__ else "utf-8")
        layout.addRow(t("import_export.encoding") + ":", self._encoding_combo)

        self._delimiter_combo = QComboBox(self)
        self._delimiter_combo.addItems([t("import_export.delimiter.comma"), t("import_export.delimiter.semicolon"), t("import_export.delimiter.tab"), t("import_export.delimiter.pipe")])
        layout.addRow(t("import_export.delimiter") + ":", self._delimiter_combo)

        self._header_check = QCheckBox(t("import_export.has_header"), self)
        self._header_check.setChecked(True)
        layout.addRow(self._header_check)

        self._batch_spin = QSpinBox(self)
        self._batch_spin.setRange(10, 10000)
        self._batch_spin.setValue(500)
        self._batch_spin.setSuffix(f" {t('import_export.rows_per_batch')}")
        layout.addRow(t("import_export.label.batch_size"), self._batch_spin)

    def encoding(self) -> str:
        return self._encoding_combo.currentText()

    def delimiter(self) -> str:
        raw = self._delimiter_combo.currentText()
        return raw[0]

    def has_header(self) -> bool:
        return self._header_check.isChecked()

    def batch_size(self) -> int:
        return self._batch_spin.value()


class ConfirmPage(QWizardPage):
    """Final page: summary and execute."""

    def __init__(self, is_import: bool) -> None:
        super().__init__()
        self._is_import = is_import
        self.setTitle(t("import_export.step_confirm"))
        layout = QVBoxLayout(self)
        self._summary = QLabel(self)
        self._summary.setWordWrap(True)
        layout.addWidget(self._summary)
        self._result = QLabel(self)
        self._result.setWordWrap(True)
        layout.addWidget(self._result)
        layout.addStretch()

    def set_summary(self, text: str) -> None:
        self._summary.setText(text)

    def set_result(self, text: str) -> None:
        self._result.setText(text)


class ImportWizard(QWizard):
    """Multi-step import wizard — CSV/Excel → Database."""

    def __init__(self, connection_id: str, database: str, table: str, parent=None) -> None:
        super().__init__(parent)
        self._connection_id = connection_id
        self._database = database
        self._table = table
        self.setWindowTitle(t("import_export.wizard_title_import"))
        self.setMinimumSize(500, 400)
        self._format_page = FormatPage(is_import=True)
        self._file_page = FilePage(is_import=True)
        self._options_page = OptionsPage()
        self._confirm_page = ConfirmPage(is_import=True)
        self.addPage(self._format_page)
        self.addPage(self._file_page)
        self.addPage(self._options_page)
        self.addPage(self._confirm_page)

    def accept(self) -> None:
        self._run_import()
        super().accept()

    def _run_import(self) -> None:
        path = self._file_page.file_path()
        fmt = self._format_page.selected_format()
        enc = self._options_page.encoding()
        delim = self._options_page.delimiter()
        has_header = self._options_page.has_header()
        batch_size = self._options_page.batch_size()

        try:
            rows = []
            if fmt == "txt":
                with open(path, "r", encoding=enc) as f:
                    lines = f.readlines()
                if has_header and lines:
                    headers = lines[0].strip().split(delim)
                    for line in lines[1:]:
                        vals = line.strip().split(delim)
                        rows.append(dict(zip(headers, vals)))
                else:
                    for line in lines:
                        rows.append({"value": line.strip()})

            elif fmt == "csv":
                with open(path, "r", encoding=enc) as f:
                    reader = csv.DictReader(f, delimiter=delim) if has_header else csv.DictReader(f, delimiter=delim, fieldnames=[])
                    rows = [dict(r) for r in reader]

            elif fmt == "json":
                import json
                with open(path, "r", encoding=enc) as f:
                    data = json.load(f)
                    rows = data if isinstance(data, list) else [data]

            elif fmt == "xml":
                import xml.etree.ElementTree as ElementTree
                tree = ElementTree.parse(path)
                root = tree.getroot()
                for row_elem in root.findall("row"):
                    row = {}
                    for child in row_elem:
                        row[child.tag] = child.text
                    rows.append(row)

            elif fmt == "html":
                from html.parser import HTMLParser
                class TableParser(HTMLParser):
                    def __init__(self):
                        super().__init__()
                        self.headers = []
                        self.rows = []
                        self._in_th = False
                        self._in_td = False
                        self._current_row = []
                        self._is_header = True
                    def handle_starttag(self, tag, attrs):
                        if tag == "th":
                            self._in_th = True
                        elif tag == "td":
                            self._in_td = True
                        elif tag == "tr":
                            self._current_row = []
                    def handle_endtag(self, tag):
                        if tag == "th":
                            self._in_th = False
                        elif tag == "td":
                            self._in_td = False
                        elif tag == "tr":
                            if self._is_header and not self.headers:
                                self.headers = self._current_row
                                self._is_header = False
                            else:
                                self.rows.append(self._current_row)
                    def handle_data(self, data):
                        if self._in_th:
                            self.headers.append(data.strip())
                        elif self._in_td:
                            self._current_row.append(data.strip())
                parser = TableParser()
                with open(path, "r", encoding=enc) as f:
                    parser.feed(f.read())
                if has_header and parser.headers:
                    for r in parser.rows:
                        rows.append(dict(zip(parser.headers, r)))
                else:
                    for r in parser.rows:
                        rows.append({str(i): v for i, v in enumerate(r)})

            elif fmt == "sql":
                with open(path, "r", encoding=enc) as f:
                    content = f.read()
                import re
                inserts = re.findall(r"INSERT\s+INTO\s+`?(\w+)`?\s*\(([^)]+)\)\s*VALUES\s*\(([^)]+)\)", content, re.IGNORECASE)
                if inserts:
                    cols = [c.strip().strip("`") for c in inserts[0][1].split(",")]
                    for match in inserts:
                        vals_raw = match[2]
                        vals = [v.strip().strip("'\"") for v in vals_raw.split(",")]
                        rows.append(dict(zip(cols, vals)))

            elif fmt == "xlsx":
                from openpyxl import load_workbook
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))] if has_header else []
                start = 2 if has_header else 1
                for row in ws.iter_rows(min_row=start, values_only=True):
                    row_data = dict(zip(headers, [str(v) if v is not None else None for v in row])) if has_header else {}
                    if not has_header:
                        row_data = {str(i): str(v) if v is not None else None for i, v in enumerate(row)}
                    rows.append(row_data)

            if not rows:
                self._confirm_page.set_result(t("import_export.msg.file_empty"))
                return

            connector = connection_pool.get(self._connection_id)
            if not connector:
                self._confirm_page.set_result(t("import_export.msg.not_connected"))
                return

            total = 0
            for i in range(0, len(rows), batch_size):
                batch = rows[i:i + batch_size]
                inserted = pool_loop.run_until_complete(
                    connector.batch_insert(self._database, self._table, batch)
                )
                total += inserted

            self._confirm_page.set_result(t("import_export.import_success", count=total))
        except Exception as e:
            self._confirm_page.set_result(f"❌ 导入失败: {e}")


class ExportWizard(QWizard):
    """Multi-step export wizard — Database → CSV/Excel/JSON."""

    def __init__(self, connection_id: str, database: str, table: str, parent=None) -> None:
        super().__init__(parent)
        self._connection_id = connection_id
        self._database = database
        self._table = table
        self.setWindowTitle(t("import_export.wizard_title_export"))
        self.setMinimumSize(500, 400)
        self._format_page = FormatPage(is_import=False)
        self._file_page = FilePage(is_import=False)
        self._options_page = OptionsPage()
        self._confirm_page = ConfirmPage(is_import=False)
        self.addPage(self._format_page)
        self.addPage(self._file_page)
        self.addPage(self._options_page)
        self.addPage(self._confirm_page)

    def accept(self) -> None:
        self._run_export()
        super().accept()

    def _run_export(self) -> None:
        path = self._file_page.file_path()
        fmt = self._format_page.selected_format()
        enc = self._options_page.encoding()
        delim = self._options_page.delimiter()
        has_header = self._options_page.has_header()

        connector = connection_pool.get(self._connection_id)
        if not connector:
            self._confirm_page.set_result(t("import_export.msg.not_connected"))
            return

        try:
            sql = f"SELECT * FROM `{self._database}`.`{self._table}`"
            result = pool_loop.run_until_complete(connector.execute(sql))
            if not result.columns:
                self._confirm_page.set_result(t("import_export.msg.no_data"))
                return

            cols = [c.name for c in result.columns]

            if fmt == "txt":
                with open(path, "w", encoding=enc) as f:
                    if has_header:
                        f.write(delim.join(cols) + "\n")
                    for row in result.rows:
                        f.write(delim.join(str(v) if v is not None else "" for v in row) + "\n")

            elif fmt == "csv":
                with open(path, "w", newline="", encoding=enc) as f:
                    writer = csv.writer(f, delimiter=delim)
                    if has_header:
                        writer.writerow(cols)
                    for row in result.rows:
                        writer.writerow([str(v) if v is not None else "" for v in row])

            elif fmt == "json":
                import json
                data = [dict(zip(cols, [str(v) if v is not None else None for v in row])) for row in result.rows]
                with open(path, "w", encoding=enc) as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)

            elif fmt == "xml":
                lines = ['<?xml version="1.0" encoding="UTF-8"?>', f"<{self._table}>"]
                for row in result.rows:
                    lines.append("  <row>")
                    for c, v in zip(cols, row):
                        val = str(v) if v is not None else ""
                        lines.append(f"    <{c}>{val}</{c}>")
                    lines.append("  </row>")
                lines.append(f"</{self._table}>")
                with open(path, "w", encoding=enc) as f:
                    f.write("\n".join(lines))

            elif fmt == "html":
                lines = [
                    "<!DOCTYPE html><html><head>",
                    f"<meta charset='{enc}'>",
                    f"<title>{self._table}</title>",
                    "<style>table{border-collapse:collapse}th,td{border:1px solid #ccc;padding:4px 8px}th{background:#f5f5f5}</style>",
                    "</head><body>",
                    "<table><thead><tr>",
                ]
                if has_header:
                    lines.append("".join(f"<th>{c}</th>" for c in cols))
                lines.append("</tr></thead><tbody>")
                for row in result.rows:
                    lines.append("<tr>" + "".join(f"<td>{str(v) if v is not None else ''}</td>" for v in row) + "</tr>")
                lines.append("</tbody></table></body></html>")
                with open(path, "w", encoding=enc) as f:
                    f.write("\n".join(lines))

            elif fmt == "sql":
                lines = [f"-- Export table `{self._table}`", ""]
                for row in result.rows:
                    vals = ", ".join(f"'{str(v).replace(chr(39), chr(39)*2)}'" if v is not None else "NULL" for v in row)
                    lines.append(f"INSERT INTO `{self._table}` ({', '.join(f'`{c}`' for c in cols)}) VALUES ({vals});")
                with open(path, "w", encoding=enc) as f:
                    f.write("\n".join(lines))

            elif fmt == "xlsx":
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = self._table
                if has_header:
                    ws.append(cols)
                for row in result.rows:
                    ws.append([v for v in row])
                wb.save(path)

            self._confirm_page.set_result(t("import_export.export_success", count=len(result.rows), path=path))
        except Exception as e:
            self._confirm_page.set_result(f"❌ 导出失败: {e}")
